import argparse
import os
import sys

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# 1. 绕过可能的代理配置
os.environ["NO_PROXY"] = "*"
os.environ["HTTP_PROXY"] = ""
os.environ["HTTPS_PROXY"] = ""

sys.path.append(os.getcwd())

from stock_monitor.config.manager import ConfigManager
from stock_monitor.core.data.stock_data_fetcher import StockDataFetcher
from stock_monitor.core.engine.quant_engine import QuantEngine


def compute_indicators(df: pd.DataFrame) -> pd.DataFrame:
    """计算 K线 均线、BOLL、MACD、RSI 及成交量能指标"""
    df = df.copy()
    if df.empty or len(df) < 20:
        return df

    # 1. 价格均线
    df["MA5"] = df["close"].rolling(5).mean()
    df["MA10"] = df["close"].rolling(10).mean()
    df["MA20"] = df["close"].rolling(20).mean()
    df["MA60"] = df["close"].rolling(60).mean()

    # 2. BOLL
    try:
        df.ta.bbands(length=20, std=2, append=True)
        df = df.rename(
            columns={
                "BBL_20_2.0": "BOLL_下轨",
                "BBM_20_2.0": "BOLL_中轨",
                "BBU_20_2.0": "BOLL_上轨",
            }
        )
        if "BBB_20_2.0" in df.columns:
            df = df.drop(columns=["BBB_20_2.0"])
        if "BBP_20_2.0" in df.columns:
            df = df.drop(columns=["BBP_20_2.0"])
    except Exception as e:
        print(f"计算 BOLL 失败: {e}")
        df["BOLL_下轨"] = np.nan
        df["BOLL_中轨"] = np.nan
        df["BOLL_上轨"] = np.nan

    # 3. MACD
    try:
        df.ta.macd(fast=12, slow=26, signal=9, append=True)
        df = df.rename(
            columns={
                "MACD_12_26_9": "MACD_DIFF",
                "MACDs_12_26_9": "MACD_DEA",
                "MACDh_12_26_9": "MACD_BAR",
            }
        )
    except Exception as e:
        print(f"计算 MACD 失败: {e}")
        df["MACD_DIFF"] = np.nan
        df["MACD_DEA"] = np.nan
        df["MACD_BAR"] = np.nan

    # 4. RSI
    try:
        df.ta.rsi(length=14, append=True)
    except Exception as e:
        print(f"计算 RSI 失败: {e}")
        df["RSI_14"] = np.nan

    # 5. 成交量均线
    df["成交量_MA5"] = df["volume"].rolling(5).mean()
    df["成交量_MA20"] = df["volume"].rolling(20).mean()

    return df


def fetch_realtime_stocks():
    """获取全量 A 股实时行情"""
    import akshare as ak

    print("[抓取] 正在通过 Eastmoney 接口获取全量 A 股实时行情...")
    try:
        df = ak.stock_zh_a_spot_em()
        if df is not None and not df.empty:
            print(f"成功获取 {len(df)} 只股票的实时数据。")
            return df
    except Exception as e:
        print(f"Eastmoney 获取失败: {e}，尝试使用 Sina 接口备份...")

    try:
        df = ak.stock_zh_a_spot()
        if df is not None and not df.empty:
            column_mapping = {
                "代码": "代码",
                "名称": "名称",
                "最新价": "最新价",
                "涨跌幅": "涨跌幅",
                "昨收": "昨收",
                "今开": "今开",
                "最高": "最高",
                "最低": "最低",
                "成交量": "成交量",
                "成交额": "成交额",
            }
            df = df.rename(columns=column_mapping)
            print(f"Sina 备份接口成功获取 {len(df)} 只股票的数据。")
            return df
    except Exception as e:
        print(f"Sina 备份接口获取也失败了: {e}")

    raise RuntimeError("所有实时行情获取接口均失败，请检查网络。")


def get_watchlist() -> list[str]:
    """获取自选股"""
    cm = ConfigManager()
    watchlist = cm.get("user_stocks", [])
    if not watchlist:
        watchlist = ["sh600460", "sz000559", "sz002600"]
    return watchlist


def resolve_stock_name(fetcher: StockDataFetcher, code: str) -> str:
    try:
        name = fetcher.name_registry.get_name(code)
        if name and name != code:
            return name
    except Exception:
        pass
    indices_names = {
        "sh000001": "上证指数",
        "sz399001": "深证成指",
        "sz399006": "创业板指",
        "sh000300": "沪深300",
    }
    return indices_names.get(code, code)


def export_to_excel(
    output_path="analysis_reports/stock_export_report.xlsx",
    include_history=False,
    history_symbols=None,
    offset=350,
    mode="watchlist",
):
    """
    抓取数据并导出到 Excel 表格中，进行精美的样式美化。
    """
    # 确保保存路径的目录存在
    dir_name = os.path.dirname(output_path)
    if dir_name and not os.path.exists(dir_name):
        os.makedirs(dir_name)

    writer = pd.ExcelWriter(output_path, engine="openpyxl")

    # 模式 A: 导出全量 A 股实时行情
    if mode in ["all", "both"]:
        try:
            df_realtime = fetch_realtime_stocks()
            if "序号" in df_realtime.columns:
                df_realtime = df_realtime.drop(columns=["序号"])
            if "涨跌幅" in df_realtime.columns:
                df_realtime["涨跌幅"] = pd.to_numeric(
                    df_realtime["涨跌幅"], errors="coerce"
                )
                df_realtime = df_realtime.sort_values(by="涨跌幅", ascending=False)

            df_realtime.to_excel(writer, sheet_name="全量A股实时行情", index=False)
            print("全量 A 股实时行情已写入 Excel。")
        except Exception as e:
            print(f"获取全量实时数据失败: {e}")

    # 模式 B: 导出自选股历史数据及指标
    if include_history or mode in ["watchlist", "both"]:
        fetcher = StockDataFetcher()
        mootdx_client = fetcher.mootdx_client
        if not mootdx_client:
            print("错误：网路故障或初始化 mootdx 行情引擎失败，跳过自选股历史获取！")
        else:
            quant_engine = QuantEngine(mootdx_client)
            # 如果没有传入指定的代码，自动从配置中读取自选股
            watchlist = (
                history_symbols if history_symbols is not None else get_watchlist()
            )
            print(f"已加载自选股：{watchlist}")

            all_history_records = []
            summary_data = []

            for code in watchlist:
                print(
                    f"[抓取] 正在获取 {code} ({resolve_stock_name(fetcher, code)}) 历史K线..."
                )
                try:
                    df_raw = quant_engine.fetch_bars(code, category=9, offset=offset)
                    if df_raw.empty:
                        print(f"警告：未拉取到 {code} 的历史K线。")
                        continue

                    df_indicators = compute_indicators(df_raw)
                    if df_indicators.empty:
                        continue

                    stock_name = resolve_stock_name(fetcher, code)
                    df_indicators.insert(0, "股票代码", code)
                    df_indicators.insert(1, "股票名称", stock_name)

                    if "datetime" in df_indicators.columns:
                        df_indicators["日期"] = pd.to_datetime(
                            df_indicators["datetime"]
                        ).dt.strftime("%Y-%m-%d")
                        df_indicators = df_indicators.drop(columns=["datetime"])
                    else:
                        df_indicators["日期"] = "未知"

                    column_order = [
                        "日期",
                        "股票代码",
                        "股票名称",
                        "open",
                        "close",
                        "high",
                        "low",
                        "volume",
                        "amount",
                        "MA5",
                        "MA10",
                        "MA20",
                        "MA60",
                        "BOLL_上轨",
                        "BOLL_中轨",
                        "BOLL_下轨",
                        "MACD_DIFF",
                        "MACD_DEA",
                        "MACD_BAR",
                        "RSI_14",
                        "成交量_MA5",
                        "成交量_MA20",
                    ]
                    columns_to_keep = [
                        col for col in column_order if col in df_indicators.columns
                    ]
                    df_final = df_indicators[columns_to_keep].copy()

                    df_final = df_final.rename(
                        columns={
                            "open": "开盘价",
                            "close": "收盘价",
                            "high": "最高价",
                            "low": "最低价",
                            "volume": "成交量",
                            "amount": "成交额",
                        }
                    )

                    df_final = df_final.sort_values(by="日期", ascending=False)

                    # 写入个股 Sheet
                    sheet_title = f"{code}_{stock_name}"
                    if len(sheet_title) > 30:
                        sheet_title = sheet_title[:30]
                    df_final.to_excel(writer, sheet_name=sheet_title, index=False)
                    all_history_records.append(df_final)

                    # 提取最新一天数据做汇总
                    latest_row = df_final.iloc[0]
                    prev_row = df_final.iloc[1] if len(df_final) > 1 else latest_row
                    close_price = latest_row["收盘价"]
                    prev_close = prev_row["收盘价"]
                    pct_change = (
                        (close_price - prev_close) / prev_close if prev_close else 0.0
                    )

                    boll_pos = "中位震荡"
                    if pd.notna(latest_row.get("BOLL_上轨")) and pd.notna(
                        latest_row.get("BOLL_下轨")
                    ):
                        if close_price >= latest_row["BOLL_上轨"] * 0.99:
                            boll_pos = "🔥 上轨阻力"
                        elif close_price <= latest_row["BOLL_下轨"] * 1.01:
                            boll_pos = "🟢 下轨支撑"

                    macd_status = "盘整"
                    if pd.notna(latest_row.get("MACD_BAR")):
                        bar = latest_row["MACD_BAR"]
                        prev_bar = prev_row.get("MACD_BAR", 0.0)
                        if bar > 0 and prev_bar <= 0:
                            macd_status = "🔴 金叉"
                        elif bar < 0 and prev_bar >= 0:
                            macd_status = "🟢 死叉"
                        elif bar > 0:
                            macd_status = "多头运行"
                        elif bar < 0:
                            macd_status = "空头运行"

                    summary_data.append(
                        {
                            "股票代码": code,
                            "股票名称": stock_name,
                            "最新日期": latest_row["日期"],
                            "最新价": close_price,
                            "今日涨跌幅": pct_change,
                            "RSI(14)": latest_row.get("RSI_14", np.nan),
                            "MACD信号": macd_status,
                            "BOLL位置": boll_pos,
                            "成交量(手)": latest_row["成交量"] / 100.0
                            if pd.notna(latest_row["成交量"])
                            else np.nan,
                            "成交额(万)": latest_row["成交额"] / 10000.0
                            if pd.notna(latest_row["成交额"])
                            else np.nan,
                        }
                    )
                except Exception as e:
                    print(f"处理自选股 {code} 失败: {e}")

            if all_history_records:
                df_summary = pd.DataFrame(summary_data)
                df_summary = df_summary.sort_values(by="今日涨跌幅", ascending=False)
                df_summary.to_excel(writer, sheet_name="自选股技术面概览", index=False)

                df_combined = pd.concat(all_history_records, ignore_index=True)
                df_combined = df_combined.sort_values(
                    by=["日期", "股票代码"], ascending=[False, True]
                )
                df_combined.to_excel(
                    writer, sheet_name="自选股合并历史明细", index=False
                )

    writer.close()

    # 美化样式渲染
    from openpyxl import load_workbook

    wb = load_workbook(output_path)

    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    header_font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Microsoft YaHei", size=10, color="333333")
    border_side = Side(border_style="thin", color="E0E0E0")
    thin_border = Border(
        left=border_side, right=border_side, top=border_side, bottom=border_side
    )

    red_fill = PatternFill(start_color="FFD2D2", end_color="FFD2D2", fill_type="solid")
    red_font = Font(name="Microsoft YaHei", size=10, color="9C0006")
    green_fill = PatternFill(
        start_color="D2FFD2", end_color="D2FFD2", fill_type="solid"
    )
    green_font = Font(name="Microsoft YaHei", size=10, color="006100")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.row_dimensions[1].height = 28
        ws.auto_filter.ref = ws.dimensions

        for col_idx, col in enumerate(ws.columns, 1):
            col_letter = get_column_letter(col_idx)
            header_value = str(col[0].value)

            max_len = sum(
                2 if "\u4e00" <= char <= "\u9fff" else 1 for char in header_value
            )
            is_numeric = False
            if any(
                k in header_value
                for k in [
                    "价",
                    "幅",
                    "额",
                    "量",
                    "率",
                    "值",
                    "最高",
                    "最低",
                    "MA",
                    "轨",
                    "DIFF",
                    "DEA",
                    "BAR",
                    "RSI",
                    "最新价",
                ]
            ):
                is_numeric = True

            for cell in col:
                cell.border = thin_border

                if cell.row == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    continue

                cell.font = data_font

                if is_numeric:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    if "幅" in header_value or "率" in header_value:
                        try:
                            if cell.value is not None and not pd.isna(cell.value):
                                val = float(cell.value)
                                if abs(val) > 1.0:
                                    cell.value = val / 100.0
                                cell.number_format = "0.00%"

                                if cell.value > 0:
                                    cell.fill = red_fill
                                    cell.font = red_font
                                elif cell.value < 0:
                                    cell.fill = green_fill
                                    cell.font = green_font
                        except ValueError:
                            pass
                    elif "量" in header_value:
                        cell.number_format = "#,##0"
                    elif "额" in header_value or "万" in header_value:
                        cell.number_format = "#,##0.00"
                    else:
                        cell.number_format = "0.00"
                        if header_value == "MACD_BAR":
                            try:
                                if cell.value is not None and not pd.isna(cell.value):
                                    val = float(cell.value)
                                    if val > 0:
                                        cell.font = red_font
                                    elif val < 0:
                                        cell.font = green_font
                            except ValueError:
                                pass
                else:
                    if len(str(cell.value)) <= 10:
                        cell.alignment = Alignment(
                            horizontal="center", vertical="center"
                        )
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")

                    if header_value == "MACD信号":
                        if cell.value == "🔴 金叉":
                            cell.fill = red_fill
                            cell.font = red_font
                        elif cell.value == "🟢 死叉":
                            cell.fill = green_fill
                            cell.font = green_font
                    elif header_value == "BOLL位置":
                        if "阻力" in str(cell.value):
                            cell.fill = red_fill
                            cell.font = red_font
                        elif "支撑" in str(cell.value):
                            cell.fill = green_fill
                            cell.font = green_font

                if cell.value is None or pd.isna(cell.value):
                    cell_len = 0
                else:
                    if is_numeric:
                        try:
                            val = float(cell.value)
                            if "幅" in header_value or "率" in header_value:
                                cell_len = len(f"{val * 100:.2f}%")
                            elif "量" in header_value:
                                cell_len = len(f"{int(round(val)):,}")
                            elif "额" in header_value or "万" in header_value:
                                cell_len = len(f"{val:,.2f}")
                            else:
                                cell_len = len(f"{val:.2f}")
                        except (ValueError, TypeError):
                            cell_len = sum(
                                2 if "\u4e00" <= char <= "\u9fff" else 1
                                for char in str(cell.value)
                            )
                    else:
                        cell_len = sum(
                            2 if "\u4e00" <= char <= "\u9fff" else 1
                            for char in str(cell.value)
                        )

                if cell_len > max_len:
                    max_len = cell_len

            ws.column_dimensions[col_letter].width = max(12, min(45, max_len + 5))

    wb.save(output_path)
    print(f"\n[完成] 报表生成成功，路径: {output_path}")


def main():
    parser = argparse.ArgumentParser(description="股票数据导出工具")
    parser.add_argument(
        "-m",
        "--mode",
        choices=["watchlist", "all", "both"],
        default="watchlist",
        help="导出模式：watchlist(自选股指标历史数据), all(全量A股实时数据), both(两者都导出)",
    )
    parser.add_argument(
        "-o",
        "--output",
        default="analysis_reports/stock_export_report.xlsx",
        help="Excel 输出文件路径",
    )
    parser.add_argument(
        "-n",
        "--offset",
        type=int,
        default=350,
        help="自选股历史 K 线获取天数（默认350天）",
    )
    args = parser.parse_args()

    print("=========================================")
    print(f"  股票数据导出工具 | 当前模式: {args.mode}")
    print("=========================================")

    include_history = args.mode in ["watchlist", "both"]

    export_to_excel(
        output_path=args.output,
        include_history=include_history,
        history_symbols=None,
        offset=args.offset,
        mode=args.mode,
    )


if __name__ == "__main__":
    main()
