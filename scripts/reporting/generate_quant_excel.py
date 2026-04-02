import json
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def generate_report():
    # 1. 加载数据
    with open("analysis_reports/rsrs_backtest_results.json", encoding="utf-8") as f:
        rsrs_data = json.load(f)
    with open("analysis_reports/confluence_results.json", encoding="utf-8") as f:
        conf_data = json.load(f)

    # 提取 Markdown 中的基本评分 (Regex)
    with open("analysis_reports/stock_analysis_report.md", encoding="utf-8") as f:
        md_content = f.read()

    stocks = [
        {"name": "万向钱潮", "code": "000559"},
        {"name": "领益智造", "code": "002600"},
        {"name": "多氟多", "code": "002407"},
        {"name": "士兰微", "code": "600460"},
    ]

    # 正则提取评分等信息
    for s in stocks:
        pattern = rf"\| \*\*{s['name']}\*\* \| {s['code']} \| ([\-\d]+) \| .* \| .* \| ([\d\-\.\+ ]+) \(.*\) \| ([🟢🟡🔴].*?) \|"
        match = re.search(pattern, md_content)
        if match:
            s["score"] = int(match.group(1))
            s["rsrs_z"] = match.group(2).strip()
            s["audit"] = match.group(3).strip()
        else:
            s["score"], s["rsrs_z"], s["audit"] = 0, "0.0", "未知"

    wb = Workbook()

    # --- Sheet 1: 综合仪表盘 ---
    ws1 = wb.active
    ws1.title = "综合仪表盘"

    headers = [
        "股票名称",
        "代码",
        "量化评分",
        "RSRS标准分",
        "审计状态",
        "策略共振胜率",
        "建议",
    ]
    ws1.append(headers)

    # 样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )

    for cell in ws1[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for s in stocks:
        conf_wr = conf_data.get(f"{s['code']}_confluence_60m", {}).get("win_rate", 0)
        advice = "观望"
        if s["score"] >= 3:
            advice = "重点关注"
        if conf_wr > 0.7:
            advice = "双因子共振买点"

        ws1.append(
            [
                s["name"],
                s["code"],
                s["score"],
                s["rsrs_z"],
                s["audit"],
                f"{conf_wr*100:.1f}%",
                advice,
            ]
        )

    # --- Sheet 2: RSRS 专项回测 ---
    ws2 = wb.create_sheet("RSRS 专项回测")
    ws2.append(["股票", "代码", "周期", "信号数", "历史胜率", "平均利润"])

    for s in stocks:
        for tf, cat in [("日线", "daily"), ("60分钟", "60m")]:
            k = f"{s['code']}_{cat}"
            data = rsrs_data.get(k, {})
            ws2.append(
                [
                    s["name"],
                    s["code"],
                    tf,
                    data.get("total_signals", 0),
                    f"{data.get('win_rate', 0)*100:.1f}%",
                    f"{data.get('avg_profit', 0)*100:.2f}%",
                ]
            )

    # --- Sheet 3: 策略共振详情 ---
    ws3 = wb.create_sheet("策略共振回测")
    ws3.append(["股票", "代码", "周期", "共振频次", "极致胜率", "平均反弹幅度"])
    for s in stocks:
        for tf, key_suffix in [("日线", "daily"), ("60分钟", "60m")]:
            k = f"{s['code']}_confluence_{key_suffix}"
            data = conf_data.get(k, {})
            ws3.append(
                [
                    s["name"],
                    s["code"],
                    tf,
                    data.get("total_signals", 0),
                    f"{data.get('win_rate', 0)*100:.1f}%",
                    f"{data.get('avg_profit', 0)*100:.2f}%",
                ]
            )

    # 列宽调整
    for sheet in wb.worksheets:
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            sheet.column_dimensions[column_letter].width = max_length + 4

    wb.save("analysis_reports/stock_quant_report.xlsx")
    print("Excel Quant Report Generated.")


if __name__ == "__main__":
    generate_report()
