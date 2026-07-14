"""
暗盘资金数据 Excel/CSV 导出模块
收盘后生成：
  Sheet 1 "全市场暗盘"  —— 全市场5000+只股票，含明盘+暗盘数据，可直接筛选
  Sheet 2 "自选股暗盘"  —— 自选股子集，高亮展示
  CSV 文件 —— 全市场暗盘数据，便于其他工具处理
"""

from __future__ import annotations

import csv
from datetime import datetime, timedelta
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from stock_monitor.services.dark_trade.service import fetch_all_dark_trade
from stock_monitor.utils.logger import app_logger

# ── 明盘行情API（东方财富市场实时数据）──────────────────────────────────────
_QUOTE_URL = "https://push2.eastmoney.com/api/qt/clist/get"
_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Referer": "https://quote.eastmoney.com/",
}


def fetch_market_quotes_all() -> dict[str, dict]:
    """
    批量获取全市场 A 股明盘行情（收盘价、涨跌幅、成交量、成交额）
    返回: {code(6位): {close, pct_chg, volume, amount, ...}}
    """
    result: dict[str, dict] = {}
    # 分两个市场：上海(1) + 深圳(0)
    for fs in ["m:1+t:2,m:1+t:23", "m:0+t:6,m:0+t:80,m:0+t:81"]:
        pn = 1
        while True:
            params = {
                "pn": pn,
                "pz": 1000,
                "po": 1,
                "np": 1,
                "ut": "bd1d9ddb04089700cf9c27f6f7426281",
                "fltt": 2,
                "invt": 2,
                "fid": "f3",
                "fs": fs,
                "fields": "f12,f14,f2,f3,f5,f6",
                # f12=代码 f14=名称 f2=收盘价 f3=涨跌幅 f5=成交量(手) f6=成交额
            }
            try:
                resp = requests.get(
                    _QUOTE_URL, params=params, headers=_HEADERS, timeout=12
                )
                data = resp.json().get("data", {}) or {}
                diff = data.get("diff", [])
            except Exception as e:
                app_logger.warning(f"[DarkExport] 获取明盘行情第{pn}页失败: {e}")
                break

            for item in diff:
                code = str(item.get("f12", "")).zfill(6)
                if code:
                    result[code] = {
                        "name": item.get("f14", ""),
                        "close": item.get("f2", 0),
                        "pct_chg": item.get("f3", 0),
                        "volume": item.get("f5", 0),  # 手
                        "amount": item.get("f6", 0),  # 元
                    }

            if len(diff) < 1000:
                break
            pn += 1
            if pn > 20:
                break

    app_logger.info(f"[DarkExport] 明盘行情获取完成: {len(result)} 只股票")
    return result


def _get_recent_trade_dates(n: int = 5) -> list[str]:
    """获取最近N个交易日日期列表（简单跳过周末）"""
    dates: list[str] = []
    current = datetime.now()
    while len(dates) < n:
        if current.weekday() < 5:
            dates.append(current.strftime("%Y%m%d"))
        current -= timedelta(days=1)
    return dates  # 最新的在前


def _make_header_style(wb: Workbook):
    """创建表头样式"""
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(
        start_color="1F4E79", end_color="1F4E79", fill_type="solid"
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    return header_font, header_fill, center_align


def _auto_col_width(ws) -> None:
    """自动调整列宽"""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                clen = sum(
                    2 if "\u4e00" <= c <= "\u9fff" else 1 for c in str(cell.value)
                )
                max_len = max(max_len, clen)
        ws.column_dimensions[col_letter].width = max(10, min(40, max_len + 3))


def _apply_conditional_color(cell, value, field: str = "net") -> None:
    """为资金数值格设置红/绿颜色"""
    if value is None:
        return
    if field == "net":
        if value > 0:
            cell.font = Font(color="CC0000", bold=True)  # 红 = 流入
        elif value < 0:
            cell.font = Font(color="008000", bold=True)  # 绿 = 流出
    elif field == "pct":
        if value > 0:
            cell.font = Font(color="CC0000")
        elif value < 0:
            cell.font = Font(color="008000")


def export_dark_trade_csv(
    watchlist_codes: list,
    output_path=None,
    history_days: int = 5,
) -> Path:
    """
    收盘后导出暗盘数据CSV

    Args:
        watchlist_codes: 自选股代码列表（原始格式，如 'sh600519' / '000559'）
        output_path:    输出文件路径，默认 analysis_reports/dark_trade_YYYYMMDD.csv
        history_days:   历史天数（近N天的净流入数据列）

    Returns:
        实际保存的 Path 对象
    """

    today_str = datetime.now().strftime("%Y%m%d")

    if output_path is None:
        out_dir = Path("analysis_reports")
        out_dir.mkdir(parents=True, exist_ok=True)
        output_path = out_dir / f"dark_trade_{today_str}.csv"
    else:
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

    app_logger.info("[DarkExport] 开始生成暗盘CSV报表...")

    # ── 1. 抓取今日全量暗盘数据 ──────────────────────────────────────────────
    today_records = fetch_all_dark_trade(today_str)
    app_logger.info(f"[DarkExport] 今日暗盘记录: {len(today_records)} 条")

    # ── 2. 抓取近N-1天历史数据（今天已有，再取前N-1天）───────────────────────
    recent_dates = _get_recent_trade_dates(history_days)  # [今天, 昨天, ...]
    history_records: dict[str, list[dict]] = {today_str: today_records}
    for d in recent_dates[1:]:
        try:
            recs = fetch_all_dark_trade(d)
            history_records[d] = recs
            app_logger.info(f"[DarkExport] {d} 历史暗盘: {len(recs)} 条")
        except Exception as e:
            app_logger.warning(f"[DarkExport] {d} 历史抓取失败: {e}")
            history_records[d] = []

    # ── 3. 获取今日明盘行情 ──────────────────────────────────────────────────
    app_logger.info("[DarkExport] 抓取明盘行情...")
    quote_map = fetch_market_quotes_all()

    # ── 4. 构建完整数据行 ────────────────────────────────────────────────────
    # 以今日暗盘为主表，补充其他字段
    # 历史净流入索引 {date: {code: net_wan}}
    history_net: dict[str, dict[str, float]] = {}
    for d, recs in history_records.items():
        idx: dict[str, float] = {}
        for r in recs:
            code = r.get("4", "")
            val = r.get("6", 0)  # 暗盘净流入 = field "6"
            if code:
                try:
                    idx[code] = float(val) / 10000
                except Exception:
                    idx[code] = 0.0
        history_net[d] = idx

    def _build_row(r: dict) -> dict:
        """将单条暗盘记录扩展为完整数据行"""
        code = r.get("4", "")
        market_val = r.get("3", 0)
        market_pfx = "SH" if market_val == 1 else "SZ"
        q = quote_map.get(code, {})

        # 暗盘净流入 = field "6"（万元）
        dark_net = r.get("6", 0)
        try:
            dark_net_wan = float(dark_net) / 10000
        except Exception:
            dark_net_wan = 0.0

        # 明盘净流入 = field "7"（万元）
        regular_net = r.get("7", 0)
        try:
            regular_net_wan = float(regular_net) / 10000
        except Exception:
            regular_net_wan = 0.0

        # 主力净流入合计 = field "8" = 暗盘 + 明盘（万元）
        total_net = r.get("8", 0)
        try:
            total_net_wan = float(total_net) / 10000
        except Exception:
            total_net_wan = 0.0

        try:
            activity = float(r.get("11", 0))
        except Exception:
            activity = 0.0

        try:
            turnover = float(r.get("14", 0))
        except Exception:
            turnover = 0.0

        hist_nets = []
        for d in recent_dates:
            hist_nets.append(history_net.get(d, {}).get(code, None))

        # 连续方向天数（从最近日期起，连续正值或负值天数）
        consecutive = 0
        if hist_nets and hist_nets[0] is not None:
            if hist_nets[0] > 0:
                # 连续流入
                for v in hist_nets:
                    if v is not None and v > 0:
                        consecutive += 1
                    else:
                        break
            elif hist_nets[0] < 0:
                # 连续流出（存为负数表示流出）
                for v in hist_nets:
                    if v is not None and v < 0:
                        consecutive -= 1
                    else:
                        break

        return {
            "code": code,
            "market": market_pfx,
            "name": q.get("name", r.get("16", "")),
            "close": q.get("close", ""),
            "pct_chg": q.get("pct_chg", ""),
            "volume_wan": q.get("volume", 0) / 100
            if q.get("volume")
            else "",  # 手→万股
            "amount_yi": q.get("amount", 0) / 1e8 if q.get("amount") else "",  # 元→亿
            "dark_net": dark_net_wan,
            "regular_net": regular_net_wan,
            "total_net": total_net_wan,
            "activity": activity,
            "turnover": turnover,
            "sector1": r.get("17", ""),
            "sector2": r.get("18", ""),
            "hist_nets": hist_nets,
            "consecutive": consecutive,
        }

    all_rows = [_build_row(r) for r in today_records]

    # 清理自选股代码（去市场前缀，统一6位）
    def clean_code(c: str) -> str:
        for prefix in ("sh", "sz", "hk", "SH", "SZ", "HK"):
            if c.startswith(prefix):
                return c[len(prefix) :]
        return c

    watchlist_clean = {clean_code(c) for c in watchlist_codes}
    watchlist_rows = [row for row in all_rows if row["code"] in watchlist_clean]

    # ── 5. 写 CSV ──────────────────────────────────────────────────────────
    # 历史日期列名
    date_cols = []
    for d in recent_dates:
        dt = datetime.strptime(d, "%Y%m%d")
        date_cols.append(f"{dt.month}/{dt.day}净流入(万)")

    headers = [
        "代码",
        "市场",
        "名称",
        "收盘价",
        "涨跌幅%",
        "成交量(万股)",
        "成交额(亿)",
        "暗盘净流入(万)",
        "明盘净流入(万)",
        "主力净流入合计(万)",
        "暗盘活跃度",
        "换手率%",
        "板块1",
        "板块2",
        "连续流入天数",
    ] + date_cols

    with open(output_path, "w", newline="", encoding="utf-8-sig") as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(headers)

        for row in all_rows:
            hist = row["hist_nets"]
            data_row = [
                row["code"],
                row["market"],
                row["name"],
                row["close"] if row["close"] != "" else "",
                row["pct_chg"] if row["pct_chg"] != "" else "",
                round(float(row["volume_wan"]), 2) if row["volume_wan"] != "" else "",
                round(float(row["amount_yi"]), 4) if row["amount_yi"] != "" else "",
                round(row["dark_net"], 2),
                round(row["regular_net"], 2),
                round(row["total_net"], 2),
                round(row["activity"], 4),
                round(row["turnover"] * 100, 2) if row["turnover"] is not None else "",
                row["sector1"],
                row["sector2"],
                row["consecutive"],
            ] + [round(v, 2) if v is not None else "" for v in hist]

            writer.writerow(data_row)

    app_logger.info(f"[DarkExport] CSV已保存: {output_path}")
    return output_path

    all_rows = [_build_row(r) for r in today_records]

    # 清理自选股代码（去市场前缀，统一6位）
    def clean_code(c: str) -> str:
        for prefix in ("sh", "sz", "hk", "SH", "SZ", "HK"):
            if c.startswith(prefix):
                return c[len(prefix) :]
        return c

    watchlist_clean = {clean_code(c) for c in watchlist_codes}
    watchlist_rows = [row for row in all_rows if row["code"] in watchlist_clean]

    # ── 5. 写 Excel ──────────────────────────────────────────────────────────
    wb = Workbook()
    h_font, h_fill, h_align = _make_header_style(wb)

    def _write_sheet(ws, rows: list[dict], title: str):
        ws.title = title

        # 历史日期列名
        date_cols = []
        for d in recent_dates:
            dt = datetime.strptime(d, "%Y%m%d")
            date_cols.append(f"{dt.month}/{dt.day}净流入(万)")

        headers = [
            "代码",
            "市场",
            "名称",
            "收盘价",
            "涨跌幅%",
            "成交量(万股)",
            "成交额(亿)",
            "暗盘净流入(万)",
            "明盘净流入(万)",
            "主力净流入合计(万)",
            "暗盘活跃度",
            "换手率%",
            "板块1",
            "板块2",
            "连续流入天数",
        ] + date_cols

        ws.append(headers)

        # 表头样式
        for cell in ws[1]:
            cell.font = h_font
            cell.fill = h_fill
            cell.alignment = h_align

        for row in rows:
            hist = row["hist_nets"]
            data_row = [
                row["code"],
                row["market"],
                row["name"],
                row["close"] if row["close"] != "" else None,
                row["pct_chg"] if row["pct_chg"] != "" else None,
                round(float(row["volume_wan"]), 2) if row["volume_wan"] != "" else None,
                round(float(row["amount_yi"]), 4) if row["amount_yi"] != "" else None,
                round(row["dark_net"], 2),
                round(row["regular_net"], 2),
                round(row["total_net"], 2),
                round(row["activity"], 4),
                round(row["turnover"] * 100, 2)
                if row["turnover"] is not None
                else None,
                row["sector1"],
                row["sector2"],
                row["consecutive"],
            ] + [round(v, 2) if v is not None else None for v in hist]

            ws.append(data_row)
            excel_row = ws.max_row

            # 涨跌幅颜色
            pct_cell = ws.cell(row=excel_row, column=5)
            try:
                pct_val = float(row["pct_chg"]) if row["pct_chg"] != "" else None
                _apply_conditional_color(pct_cell, pct_val, "pct")
            except Exception:
                pass

            # 暗盘净流入颜色
            net_cell = ws.cell(row=excel_row, column=8)
            _apply_conditional_color(net_cell, row["dark_net"], "net")

            # 主力净流入合计颜色
            total_cell = ws.cell(row=excel_row, column=10)
            _apply_conditional_color(total_cell, row["total_net"], "net")

            # 历史净流入颜色
            for i, v in enumerate(hist):
                hcell = ws.cell(row=excel_row, column=16 + i)
                if v is not None:
                    _apply_conditional_color(hcell, v, "net")

        # 设置数字格式
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, float):
                    cell.alignment = Alignment(horizontal="right")

        _auto_col_width(ws)

        # 冻结首列+首行
        ws.freeze_panes = "D2"

    # Sheet 1: 全市场
    ws_all = wb.active
    _write_sheet(ws_all, all_rows, "全市场暗盘")

    # Sheet 2: 自选股
    ws_watch = wb.create_sheet("自选股暗盘")
    _write_sheet(ws_watch, watchlist_rows, "自选股暗盘")

    wb.save(str(output_path))
    app_logger.info(f"[DarkExport] Excel已保存: {output_path}")
    return output_path
